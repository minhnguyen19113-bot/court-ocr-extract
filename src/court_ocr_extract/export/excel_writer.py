from __future__ import annotations

from pathlib import Path

from court_ocr_extract.models import ExtractionResult, Participant


EXCEL_HEADERS = [
    "LOẠI ÁN",
    "SỐ THỤ LÝ",
    "NGÀY THỤ LÝ (DD/MM/YYYY)",
    "QUAN HỆ PHÁP LUẬT",
    "TƯ CÁCH TỐ TỤNG",
    "HỌ TÊN ĐƯƠNG SỰ",
    "NĂM SINH",
    "CCCD",
    "ĐỊA CHỈ",
    "HỌ TÊN CHỦ TỌA",
    "GHI CHÚ",
]


def rows_from_result(result: ExtractionResult) -> list[dict[str, str | None]]:
    participants = result.participants or [
        Participant(ghi_chu="Không nhận diện được người tham gia tố tụng")
    ]
    rows: list[dict[str, str | None]] = []
    common_note = "; ".join(result.warnings)
    for participant in participants:
        note = "; ".join(item for item in [participant.ghi_chu, common_note] if item)
        rows.append(
            {
                "LOẠI ÁN": result.case_info.loai_an,
                "SỐ THỤ LÝ": result.case_info.so_thu_ly,
                "NGÀY THỤ LÝ (DD/MM/YYYY)": result.case_info.ngay_thu_ly,
                "QUAN HỆ PHÁP LUẬT": result.case_info.quan_he_phap_luat,
                "TƯ CÁCH TỐ TỤNG": participant.tu_cach_to_tung,
                "HỌ TÊN ĐƯƠNG SỰ": participant.ho_ten,
                "NĂM SINH": participant.nam_sinh,
                "CCCD": participant.cccd,
                "ĐỊA CHỈ": participant.dia_chi,
                "HỌ TÊN CHỦ TỌA": result.case_info.chu_toa,
                "GHI CHÚ": note or None,
            }
        )
    return rows


def write_excel(result: ExtractionResult, output_path: str | Path) -> Path:
    return write_excel_from_results([result], output_path)


def write_excel_from_results(results: list[ExtractionResult], output_path: str | Path) -> Path:
    from openpyxl import Workbook

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Trich xuat"

    worksheet.append(EXCEL_HEADERS)
    for result in results:
        for row in rows_from_result(result):
            worksheet.append([row.get(header) for header in EXCEL_HEADERS])

    _format_worksheet(worksheet)
    workbook.save(output_path)
    return output_path


def _format_worksheet(worksheet) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="000000")
    wrap = Alignment(wrap_text=True, vertical="top")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {
        "A": 16,
        "B": 20,
        "C": 22,
        "D": 32,
        "E": 28,
        "F": 28,
        "G": 12,
        "H": 18,
        "I": 48,
        "J": 28,
        "K": 44,
    }
    for column_index in range(1, len(EXCEL_HEADERS) + 1):
        letter = get_column_letter(column_index)
        worksheet.column_dimensions[letter].width = widths.get(letter, 18)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
