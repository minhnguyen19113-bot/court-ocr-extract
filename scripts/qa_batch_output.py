from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any


CASE_FIELDS = {
    "loai_an": "LOAI_AN",
    "so_thu_ly": "SO_THU_LY",
    "ngay_thu_ly": "NGAY_THU_LY",
    "quan_he_phap_luat": "QUAN_HE_PHAP_LUAT",
    "chu_toa": "CHU_TOA",
}

PARTICIPANT_FIELDS = {
    "tu_cach_to_tung": "TU_CACH_TO_TUNG",
    "ho_ten": "HO_TEN",
    "nam_sinh": "NAM_SINH",
    "cccd": "CCCD",
    "dia_chi": "DIA_CHI",
}

SUSPICIOUS_NAME_RE = re.compile(
    r"\b(?:co mat|vang mat|phien toa|tai phien|duoc trieu tap|bi bat|tam giam)\b",
    re.IGNORECASE,
)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Print non-sensitive QA statistics for batch_summary.json and optional Excel."
    )
    parser.add_argument("--summary", type=Path, required=True)
    parser.add_argument("--excel", type=Path, default=None)
    args = parser.parse_args()

    summary = json.loads(args.summary.read_text(encoding="utf-8-sig"))
    report = summarize_batch(summary)
    if args.excel:
        report["excel"] = summarize_excel(args.excel)

    print(json.dumps(report, ensure_ascii=False, indent=2))


def summarize_batch(summary: dict[str, Any]) -> dict[str, Any]:
    results = [item for item in summary.get("results") or [] if isinstance(item, dict)]
    participants = [
        participant
        for result in results
        for participant in (result.get("participants") or [])
        if isinstance(participant, dict)
    ]

    report: dict[str, Any] = {
        "batch": {
            "total": summary.get("total", 0),
            "success": summary.get("success", 0),
            "failed": summary.get("failed", 0),
            "skipped": summary.get("skipped", 0),
            "result_objects": len(results),
            "participant_rows": len(participants),
        },
        "marker": {
            "found": sum(1 for result in results if result.get("marker_found") is True),
            "not_found": sum(1 for result in results if result.get("marker_found") is not True),
        },
        "review": {
            "required": sum(1 for result in results if _metadata(result).get("review_required")),
            "not_required": sum(1 for result in results if not _metadata(result).get("review_required")),
        },
        "extractors": dict(_extractor_counts(results)),
        "case_fields": _field_stats(results, CASE_FIELDS, parent_key="case_info"),
        "participant_fields": _participant_field_stats(participants),
        "warnings": dict(_warning_categories(results)),
        "suspicious_counts": _suspicious_counts(participants),
    }
    return report


def summarize_excel(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return {"error": "openpyxl_not_installed"}

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {"rows": 0, "columns": 0}
    headers = [str(item or "").strip() for item in rows[0]]
    data_rows = rows[1:]
    blank_by_column = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        blanks = sum(1 for row in data_rows if index >= len(row) or _empty(row[index]))
        blank_by_column[_safe_header(header)] = blanks
    return {
        "rows": len(data_rows),
        "columns": len(headers),
        "blank_cells_by_column": blank_by_column,
    }


def _field_stats(results: list[dict[str, Any]], fields: dict[str, str], *, parent_key: str) -> dict[str, Any]:
    stats = {}
    for field, label in fields.items():
        values = [(result.get(parent_key) or {}).get(field) for result in results]
        stats[label] = _value_stats(values)
    return stats


def _participant_field_stats(participants: list[dict[str, Any]]) -> dict[str, Any]:
    stats = {}
    for field, label in PARTICIPANT_FIELDS.items():
        values = [participant.get(field) for participant in participants]
        stats[label] = _value_stats(values)
    return stats


def _value_stats(values: list[Any]) -> dict[str, int]:
    total = len(values)
    non_empty = sum(1 for value in values if not _empty(value))
    return {
        "total": total,
        "non_empty": non_empty,
        "empty": total - non_empty,
    }


def _extractor_counts(results: list[dict[str, Any]]) -> Counter:
    counter: Counter = Counter()
    for result in results:
        metadata = _metadata(result)
        method = metadata.get("primary_extractor") or "unknown"
        counter[str(method)] += 1
    return counter


def _warning_categories(results: list[dict[str, Any]]) -> Counter:
    counter: Counter = Counter()
    for result in results:
        for warning in result.get("warnings") or []:
            category = _warning_category(str(warning))
            counter[category] += 1
    return counter


def _warning_category(value: str) -> str:
    normalized = _ascii_key(value)
    if "khong tim thay marker" in normalized:
        return "missing_marker"
    if "thieu thong tin chung" in normalized:
        return "missing_case_fields"
    if "thieu" in normalized:
        return "missing_participant_fields"
    if "confidence thap" in normalized or "can review" in normalized:
        return "needs_review_or_low_confidence"
    if "local llm" in normalized:
        return "local_llm"
    return "other"


def _suspicious_counts(participants: list[dict[str, Any]]) -> dict[str, int]:
    suspicious_names = 0
    invalid_birth_years = 0
    invalid_identity_numbers = 0
    for participant in participants:
        name_key = _ascii_key(participant.get("ho_ten"))
        if name_key and SUSPICIOUS_NAME_RE.search(name_key):
            suspicious_names += 1
        birth_year = str(participant.get("nam_sinh") or "").strip()
        if birth_year and not re.fullmatch(r"\d{4}", birth_year):
            invalid_birth_years += 1
        identity_digits = re.sub(r"\D", "", str(participant.get("cccd") or ""))
        if identity_digits and len(identity_digits) not in {9, 12}:
            invalid_identity_numbers += 1
    return {
        "suspicious_name_like_phrase": suspicious_names,
        "invalid_birth_year": invalid_birth_years,
        "invalid_identity_number_length": invalid_identity_numbers,
    }


def _metadata(result: dict[str, Any]) -> dict[str, Any]:
    metadata = result.get("metadata")
    return metadata if isinstance(metadata, dict) else {}


def _empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _safe_header(value: str) -> str:
    key = _ascii_key(value).upper()
    return re.sub(r"[^A-Z0-9]+", "_", key).strip("_") or "COLUMN"


def _ascii_key(value: Any) -> str:
    import unicodedata

    text = str(value or "").casefold()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^\w\s]", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


if __name__ == "__main__":
    main()
